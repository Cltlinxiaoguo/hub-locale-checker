# -*- coding: utf-8 -*-
"""按 config.json 将 *_zh_en.json 导出为与标准表同格式的 xlsx（文件名可带时间戳）。"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

from i18n_config import (
    get_paths,
    load_config,
    output_json_path,
    output_xlsx_path,
    resolve_js_override,
)


def _parse_cli() -> tuple[Path | None, str | None]:
    config_path = None
    args: list[str] = []
    i = 1
    while i < len(sys.argv):
        if sys.argv[i] in ("-c", "--config") and i + 1 < len(sys.argv):
            config_path = sys.argv[i + 1]
            i += 2
        else:
            args.append(sys.argv[i])
            i += 1
    js_arg = args[0] if args else None
    return (Path(config_path) if config_path else None, js_arg)


def read_reference_headers(path: Path) -> tuple[str, str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    row1 = next(ws.iter_rows(values_only=True))
    wb.close()
    zh_h = str(row1[0]).strip() if row1[0] else "中文"
    en_h = str(row1[1]).strip() if len(row1) > 1 and row1[1] else "英文"
    return zh_h, en_h


def main() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except (AttributeError, OSError):
        pass

    cfg_path, js_cli = _parse_cli()
    cfg = load_config(cfg_path)
    paths = get_paths(cfg)

    js_path = paths["js_file"]
    if js_cli:
        js_path = resolve_js_override(js_cli, cfg)

    js_json = output_json_path(js_path, cfg)
    out_xlsx = output_xlsx_path(js_path, cfg)
    ref = paths["standard_xlsx"]

    if not js_json.is_file():
        raise SystemExit(
            f"缺少 {js_json}，请先执行: python extract_compare.py\n"
            f"（或在 config.json 中设置 paths.js_file）"
        )

    zh_h, en_h = read_reference_headers(ref) if ref and ref.is_file() else ("中文", "英文")

    data = json.loads(js_json.read_text(encoding="utf-8"))
    pairs = sorted(data["pairsByPath"], key=lambda x: x["path"])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "JS提取"

    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)

    ws.append([zh_h, en_h])
    for c in ws[1]:
        c.font = header_font
        c.alignment = header_align

    for item in pairs:
        ws.append([item["zh"], item["en"]])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.alignment = cell_align

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 52

    wb.save(out_xlsx)
    print(f"配置: {cfg_path or 'config.json'}")
    print(f"Wrote {out_xlsx} ({len(pairs)} rows + header)")


if __name__ == "__main__":
    main()
